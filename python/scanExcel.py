import openpyxl
import re
import os
import requests
import concurrent.futures

# 使用绝对路径读取Excel文件
file_path = 'C:\\Users\\Administrator\\Desktop\\CODES\\HIVE-sol\\python\\tron.xlsx'

# 检查文件是否存在和格式是否正确
if not os.path.exists(file_path):
    raise FileNotFoundError(f"The file {file_path} does not exist。")
if not file_path.endswith('.xlsx'):
    raise ValueError("The file is not an .xlsx file。")

wb = openpyxl.load_workbook(file_path)
ws = wb.active

# OKLink API设置
api_key = 'd8debf59-e8a1-46da-a923-58073fa173f0'  # 替换为您的OKLink API密钥
base_url = 'https://www.oklink.com/api/v5/explorer/'

# Tron地址的正则表达式
tron_address_pattern = re.compile(r"^T[1-9A-HJ-NP-Za-km-z]{33}$")

# TRC20 USDT 合约地址
usdt_contract_address = "TR7NHqjeKQxGTCi8q8ZY4pL8otSzgjLj6t"

# 获取TRX余额的函数
def get_trx_balance(address):
    url = f"{base_url}balance?chainShortName=TRX&address={address}"
    headers = {
        'Ok-Access-Key': api_key
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        data = response.json()
        if 'data' in data and len(data['data']) > 0:
            return float(data['data'][0].get('balance', 0)) / 1_000_000  # 转换为TRX
    return 0

# 获取USDT余额的函数
def get_usdt_balance(address):
    url = f"{base_url}trc20-balance?chainShortName=TRX&address={address}&contractAddress={usdt_contract_address}"
    headers = {
        'Ok-Access-Key': api_key
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        data = response.json()
        if 'data' in data:
            for token in data['data']:
                if token['contractAddress'] == usdt_contract_address:
                    return float(token.get('balance', 0)) / 1_000_000  # 转换为USDT
    return 0

# 获取地址的TRX和USDT余额
def fetch_balances(address):
    if address and tron_address_pattern.match(address):
        try:
            trx_balance = get_trx_balance(address)
            usdt_balance = get_usdt_balance(address)
            return (address, trx_balance, usdt_balance)
        except Exception as e:
            print(f"Error fetching data for address {address}: {e}")
            return (address, 'Error', 'Error')
    else:
        return (address, 'Invalid Address', 'Invalid Address')

# 遍历每一个地址并获取USDT和TRX余额
addresses = [row[0].value for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1)]

with concurrent.futures.ThreadPoolExecutor(max_workers=50) as executor:
    results = list(executor.map(fetch_balances, addresses))

# 将结果写入Excel文件
for index, result in enumerate(results, start=2):
    address, trx_balance, usdt_balance = result
    ws.cell(row=index, column=2, value=usdt_balance)
    ws.cell(row=index, column=3, value=trx_balance)
    print(f"Processing row {index}")

# 保存修改后的Excel文件
wb.save(file_path)
print("Balances updated successfully.")
